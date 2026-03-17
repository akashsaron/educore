"""
apps/core/exports.py

Excel (.xlsx) export endpoints using openpyxl.
All exports are streamed directly as file downloads.

Endpoints:
  GET /api/exports/students/          ?section=&grade=&is_active=
  GET /api/exports/attendance/        ?section=&from_date=&to_date=
  GET /api/exports/fee-collection/    ?from_date=&to_date=
  GET /api/exports/exam-results/      ?exam_id=
"""
from io import BytesIO
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style helpers ──────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', fgColor='1a2744')
ACCENT_FILL   = PatternFill('solid', fgColor='4f8ef7')
ALT_FILL      = PatternFill('solid', fgColor='f4f6fb')
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT    = Font(bold=True, color='1a2744', size=14)
NORMAL_FONT   = Font(size=10)
CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT          = Alignment(horizontal='left',   vertical='center')
THIN          = Side(style='thin', color='e2e8f0')
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _add_header_row(ws, columns, row=1):
    for col_idx, (header, width) in enumerate(columns, 1):
        cell              = ws.cell(row=row, column=col_idx, value=header)
        cell.font         = HEADER_FONT
        cell.fill         = HEADER_FILL
        cell.alignment    = CENTER
        cell.border       = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _data_row(ws, row_idx, values):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for col_idx, value in enumerate(values, 1):
        cell           = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = NORMAL_FONT
        cell.alignment = CENTER if isinstance(value, (int, float)) else LEFT
        cell.border    = BORDER
        if fill.fill_type:
            cell.fill  = fill


def _make_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Student export ─────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_students(request):
    from apps.students.models import Student

    qs = Student.objects.select_related('section__grade', 'academic_year').all()
    if section := request.query_params.get('section'):
        qs = qs.filter(section_id=section)
    if grade := request.query_params.get('grade'):
        qs = qs.filter(section__grade_id=grade)
    if is_active := request.query_params.get('is_active'):
        qs = qs.filter(is_active=is_active.lower() == 'true')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    ws.row_dimensions[1].height = 30

    columns = [
        ('Adm. No.',     14), ('First Name',  16), ('Last Name',   16),
        ('Class',        12), ('Gender',      10), ('DOB',         14),
        ('Roll No.',     10), ('Father Name', 20), ('Phone',       16),
        ('Email',        26), ('Blood Group', 12), ('City',        14),
        ('Admission Date', 16), ('Status',    10),
    ]
    _add_header_row(ws, columns)

    for row_idx, s in enumerate(qs, 2):
        _data_row(ws, row_idx, [
            s.admission_no, s.first_name, s.last_name,
            s.current_class, s.get_gender_display(),
            str(s.date_of_birth), s.roll_number, s.father_name,
            s.parent_phone, s.parent_email, s.blood_group or '—',
            s.city, str(s.admission_date), 'Active' if s.is_active else 'Inactive',
        ])

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return _make_response(wb, 'students_export.xlsx')


# ── Attendance export ──────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance(request):
    from apps.attendance.models import AttendanceRecord
    from apps.students.models import Student
    from apps.core.models import Section

    section_id = request.query_params.get('section')
    from_date  = request.query_params.get('from_date', str(timezone_today()))
    to_date    = request.query_params.get('to_date',   str(timezone_today()))

    qs = Student.objects.filter(is_active=True)
    if section_id:
        qs = qs.filter(section_id=section_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    columns = [
        ('Adm. No.', 14), ('Student Name', 24), ('Class', 12),
        ('Present',  10), ('Absent',       10), ('Late',  10),
        ('Total',    10), ('Percentage',   12),
    ]
    _add_header_row(ws, columns)

    for row_idx, s in enumerate(qs, 2):
        records = AttendanceRecord.objects.filter(student=s, date__range=[from_date, to_date])
        total   = records.count()
        present = records.filter(status='present').count()
        absent  = records.filter(status='absent').count()
        late    = records.filter(status='late').count()
        pct     = round((present + late) / total * 100, 1) if total else 0
        _data_row(ws, row_idx, [
            s.admission_no, s.full_name, s.current_class,
            present, absent, late, total, f'{pct}%'
        ])

    ws.freeze_panes = 'A2'
    return _make_response(wb, f'attendance_{from_date}_to_{to_date}.xlsx')


def timezone_today():
    from django.utils import timezone
    return timezone.localdate()


# ── Fee collection export ──────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_fee_collection(request):
    from apps.fees.models import FeePayment

    from_date = request.query_params.get('from_date')
    to_date   = request.query_params.get('to_date')

    qs = FeePayment.objects.select_related(
        'invoice__student', 'invoice__fee_structure__category', 'collected_by'
    ).filter(status='paid')

    if from_date: qs = qs.filter(paid_date__gte=from_date)
    if to_date:   qs = qs.filter(paid_date__lte=to_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fee Collection'

    columns = [
        ('Receipt No.',  16), ('Invoice No.',  16), ('Student',     22),
        ('Class',        12), ('Category',     18), ('Amount (₹)',  14),
        ('Method',       14), ('Date',         14), ('Collected By',20),
    ]
    _add_header_row(ws, columns)

    total_amount = 0
    for row_idx, p in enumerate(qs, 2):
        total_amount += float(p.amount)
        _data_row(ws, row_idx, [
            p.receipt_no,
            p.invoice.invoice_no,
            p.invoice.student.full_name,
            p.invoice.student.current_class,
            p.invoice.fee_structure.category.name,
            float(p.amount),
            p.get_payment_method_display(),
            str(p.paid_date),
            p.collected_by.get_full_name() if p.collected_by else '—',
        ])

    # Total row
    total_row = qs.count() + 2
    ws.cell(row=total_row, column=5, value='TOTAL').font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=6, value=total_amount).font = Font(bold=True, size=11, color='1a2744')
    ws.freeze_panes = 'A2'
    return _make_response(wb, 'fee_collection.xlsx')


# ── Exam results export ────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_exam_results(request):
    from apps.exams.models import ExamResult, Exam

    exam_id = request.query_params.get('exam_id')
    qs      = ExamResult.objects.select_related(
        'student__section__grade', 'schedule__subject', 'schedule__exam'
    )
    if exam_id:
        qs = qs.filter(schedule__exam_id=exam_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Exam Results'

    columns = [
        ('Adm. No.',  14), ('Student',   22), ('Class',    12),
        ('Subject',   20), ('Max Marks', 12), ('Obtained', 12),
        ('Grade',     10), ('Result',    10), ('Absent',   10),
    ]
    _add_header_row(ws, columns)

    for row_idx, r in enumerate(qs, 2):
        pct    = round(float(r.marks_obtained) / r.schedule.max_marks * 100, 1) if r.schedule.max_marks else 0
        result = 'ABSENT' if r.is_absent else ('PASS' if r.grade != 'F' else 'FAIL')
        _data_row(ws, row_idx, [
            r.student.admission_no,
            r.student.full_name,
            r.student.current_class,
            r.schedule.subject.name,
            r.schedule.max_marks,
            float(r.marks_obtained) if not r.is_absent else 0,
            r.grade if not r.is_absent else '—',
            result,
            'Yes' if r.is_absent else 'No',
        ])

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return _make_response(wb, 'exam_results.xlsx')
